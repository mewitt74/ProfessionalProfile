import json
import openpyxl
from datetime import datetime
from collections import defaultdict

# Load the workbook to recalculate
wb = openpyxl.load_workbook('certificates/Resume Submissions.xlsx')

# Get timeline and aggregate by month
ws_timeline = wb['Sheet1']
monthly_data = defaultdict(int)
for row in ws_timeline.iter_rows(min_row=4, values_only=True):
    if row[0] and isinstance(row[0], datetime) and row[1]:
        month_key = row[0].strftime('%b %Y')
        monthly_data[month_key] += int(row[1])

# Sort by date
sorted_months = sorted(monthly_data.items(), key=lambda x: datetime.strptime(x[0], '%b %Y'))
max_count = max([count for _, count in sorted_months]) if sorted_months else 1

# Get positions and companies
ws_data = wb['Sheet2']

# More specific role groupings based on similarity
positions_by_role = {
    'Director Level': [],
    'Technical Program Manager': [],
    'Compliance Program Manager': [],
    'Senior Manager & Principal': [],
    'GRC Manager': [],
    'Security & Risk Manager': [],
    'AI & Governance Specialist': [],
    'Compliance Analyst & Specialist': [],
    'Lead & Head Roles': [],
    'Other Roles': []
}

companies_by_industry = {
    'Cloud & Infrastructure': [],
    'Enterprise Software & SaaS': [],
    'FinTech & Financial Services': [],
    'Security & Compliance': [],
    'Healthcare & Life Sciences': [],
    'E-commerce & Consumer Tech': [],
    'Consulting & Services': [],
    'Emerging Technology': []
}

# Industry mapping
industry_map = {
    'Amazon': 'Cloud & Infrastructure', 'Amazon Prime Air': 'Cloud & Infrastructure', 'Google': 'Cloud & Infrastructure',
    'Microsoft': 'Cloud & Infrastructure', 'CoreWeave': 'Cloud & Infrastructure', 'NVIDIA': 'Cloud & Infrastructure',
    'Oracle': 'Cloud & Infrastructure', 'F5': 'Cloud & Infrastructure',
    'Atlassian': 'Enterprise Software & SaaS', 'Docusign': 'Enterprise Software & SaaS', 'Asana': 'Enterprise Software & SaaS',
    'Hubspot': 'Enterprise Software & SaaS', 'Twilio': 'Enterprise Software & SaaS', 'Avalara': 'Enterprise Software & SaaS',
    'Teradata': 'Enterprise Software & SaaS', 'VeevaSystems': 'Enterprise Software & SaaS', 'insightsoftware': 'Enterprise Software & SaaS',
    'Ncontracts': 'Enterprise Software & SaaS', 'Figma': 'Enterprise Software & SaaS', 'GrafanaLabs': 'Enterprise Software & SaaS',
    'Amplitude': 'Enterprise Software & SaaS', 'Confluent': 'Enterprise Software & SaaS', 'Vercel': 'Enterprise Software & SaaS',
    'FloQast': 'Enterprise Software & SaaS', '6sense(Growth-stageB2BSaaS)': 'Enterprise Software & SaaS', 'Airtable': 'Enterprise Software & SaaS',
    'interface.ai': 'Enterprise Software & SaaS', 'MediaAlpha': 'Enterprise Software & SaaS', 'phData': 'Enterprise Software & SaaS',
    'Kyriba': 'Enterprise Software & SaaS', 'VitalhubCorp': 'Enterprise Software & SaaS',
    'Affirm': 'FinTech & Financial Services', 'Block': 'FinTech & Financial Services', 'Stripe': 'FinTech & Financial Services',
    'Coinbase': 'FinTech & Financial Services', 'Circle': 'FinTech & Financial Services', 'Paxos': 'FinTech & Financial Services',
    'Ripple': 'FinTech & Financial Services', 'AnchorageDigital': 'FinTech & Financial Services', 'GreenDot': 'FinTech & Financial Services',
    'Upstart': 'FinTech & Financial Services', 'Wrapbook': 'FinTech & Financial Services', 'BILL': 'FinTech & Financial Services',
    'BECU': 'FinTech & Financial Services', 'Allstate': 'FinTech & Financial Services', 'Symetra': 'FinTech & Financial Services',
    'CIBC': 'FinTech & Financial Services', 'Blackstone Talent Group': 'FinTech & Financial Services',
    'Evergreen Home Loans': 'FinTech & Financial Services', 'MeridianLink': 'FinTech & Financial Services', 'Experian': 'FinTech & Financial Services',
    'CrowdStrike': 'Security & Compliance', 'Zscaler': 'Security & Compliance', 'Vanta': 'Security & Compliance',
    'SecureCodeWarrior': 'Security & Compliance', 'Zenity': 'Security & Compliance', 'Entrust': 'Security & Compliance',
    'Virtru': 'Security & Compliance', 'Forta': 'Security & Compliance',
    'Milliman(IntelliScriptDivision)': 'Healthcare & Life Sciences', 'IncludedHealth': 'Healthcare & Life Sciences',
    'Baylor Scott & White Health': 'Healthcare & Life Sciences', 'CardinalHealth': 'Healthcare & Life Sciences',
    'BeiGene(BeOne)': 'Healthcare & Life Sciences', 'GlobalCommercialJazzPharma': 'Healthcare & Life Sciences',
    'SwordHealth': 'Healthcare & Life Sciences', 'HealthInTech': 'Healthcare & Life Sciences', 'Aledade': 'Healthcare & Life Sciences',
    'UnileverPrestige': 'Healthcare & Life Sciences', 'PrecisionMedicineGroup': 'Healthcare & Life Sciences',
    'BALTGroup(Balt)': 'Healthcare & Life Sciences', 'Radformation': 'Healthcare & Life Sciences',
    'AirBnB': 'E-commerce & Consumer Tech', 'DoorDash': 'E-commerce & Consumer Tech', 'Zillow': 'E-commerce & Consumer Tech',
    'Expedia Group': 'E-commerce & Consumer Tech', 'SnapChat': 'E-commerce & Consumer Tech', 'Meta': 'E-commerce & Consumer Tech',
    'Nordstrom': 'E-commerce & Consumer Tech', 'Angi': 'E-commerce & Consumer Tech', 'RealPage': 'E-commerce & Consumer Tech',
    'Deloitte': 'Consulting & Services', 'KPMGUS': 'Consulting & Services', 'KornFerry': 'Consulting & Services',
    'WilsonSonsini': 'Consulting & Services', 'HorizontalTalent': 'Consulting & Services', 'CypressHCM': 'Consulting & Services',
    'DecisionPointCorporation': 'Consulting & Services', 'Signify Technology': 'Consulting & Services',
    'IBSS': 'Consulting & Services', 'Polestar Analytics': 'Consulting & Services',
    'D-WaveQuantum': 'Emerging Technology', 'Groq': 'Emerging Technology', 'Perplexity': 'Emerging Technology',
    'Waymo': 'Emerging Technology', 'WingAviation': 'Emerging Technology', 'Phaidra': 'Emerging Technology',
    'Zania': 'Emerging Technology', 'Mozilla': 'Emerging Technology', 'Popl': 'Emerging Technology'
}

unique_companies = set()
unique_positions = set()

for row in ws_data.iter_rows(min_row=4, values_only=True):
    if row[1] and row[2]:
        company = str(row[1]).strip()
        position = str(row[2]).strip()
        pos_lower = position.lower()
        
        unique_companies.add(company)
        unique_positions.add(position)
        
        # More specific categorization
        if any(x in pos_lower for x in ['director of', 'director,', 'director risk', 'director security', 'associate director']):
            positions_by_role['Director Level'].append(position)
        elif 'technical program manager' in pos_lower or 'technical pm' in pos_lower or ('program manager' in pos_lower and 'technical' in pos_lower):
            positions_by_role['Technical Program Manager'].append(position)
        elif any(x in pos_lower for x in ['compliance program', 'pci compliance program', 'program manager, compliance']):
            positions_by_role['Compliance Program Manager'].append(position)
        elif any(x in pos_lower for x in ['senior manager', 'senior director', 'principal', 'sr. manager', 'sr manager']):
            positions_by_role['Senior Manager & Principal'].append(position)
        elif 'grc manager' in pos_lower or 'governance, risk & compliance manager' in pos_lower or 'manager, grc' in pos_lower:
            positions_by_role['GRC Manager'].append(position)
        elif any(x in pos_lower for x in ['security', 'risk manager', 'risk & compliance', 'cybersecurity']) and 'manager' in pos_lower:
            positions_by_role['Security & Risk Manager'].append(position)
        elif any(x in pos_lower for x in ['ai ', 'governance lead', 'governance specialist', 'privacy']):
            positions_by_role['AI & Governance Specialist'].append(position)
        elif any(x in pos_lower for x in ['compliance analyst', 'compliance specialist', 'grc analyst', 'compliance advisor']):
            positions_by_role['Compliance Analyst & Specialist'].append(position)
        elif any(x in pos_lower for x in ['lead', 'head of']):
            positions_by_role['Lead & Head Roles'].append(position)
        else:
            positions_by_role['Other Roles'].append(position)
        
        # Categorize company
        if company in industry_map:
            companies_by_industry[industry_map[company]].append(company)
        else:
            companies_by_industry['Emerging Technology'].append(company)

# Remove duplicates and sort
for role in positions_by_role:
    positions_by_role[role] = sorted(set(positions_by_role[role]))

for ind in companies_by_industry:
    companies_by_industry[ind] = sorted(set(companies_by_industry[ind]))

wb.close()

# Get total submissions from Sheet1 (daily counts)
wb2 = openpyxl.load_workbook('certificates/Resume Submissions.xlsx')
ws_timeline_calc = wb2['Sheet1']
total_submissions = 0
for row in ws_timeline_calc.iter_rows(min_row=4, values_only=True):
    if row[0] and isinstance(row[0], datetime) and row[1]:
        total_submissions += int(row[1])
wb2.close()

# Calculate totals for cards
total_companies = len(unique_companies)
total_title_groups = len([role for role, positions in positions_by_role.items() if positions])

# Generate timeline HTML (monthly aggregation)
timeline_html = ""
for month, count in sorted_months:
    width = int((count / max_count) * 100)
    plural = 's' if count != 1 else ''
    timeline_html += f'''            <div class="timeline-bar">
                <div class="timeline-month">{month}</div>
                <div class="timeline-visual" style="width: {width}%;">
                    <span class="timeline-count">{count} app{plural}</span>
                </div>
            </div>
'''

# Generate role data for charts
role_data = []
for role, positions in positions_by_role.items():
    if positions:
        role_data.append({
            'category': role,
            'count': len(positions),
            'positions': positions
        })
role_data.sort(key=lambda x: x['count'], reverse=True)

# Generate positions HTML with visualizations
positions_html = ""
for item in role_data:
    role = item['category']
    count = item['count']
    positions = item['positions']
    
    # Create visual bar
    bar_width = int((count / max([r['count'] for r in role_data])) * 100)
    
    positions_html += f'''            <div class="category-visual">
                <div class="category-header">
                    <h3 class="category-name">{role}</h3>
                    <span class="category-count">{count} position{"s" if count != 1 else ""}</span>
                </div>
                <div class="category-bar-container">
                    <div class="category-bar" style="width: {bar_width}%;">
                        <span class="bar-label">{count}</span>
                    </div>
                </div>
                <details class="category-details">
                    <summary>View all {count} positions</summary>
                    <div class="positions-list">
'''
    for pos in positions:
        positions_html += f'                        <div class="position-item">• {pos}</div>\n'
    positions_html += '''                    </div>
                </details>
            </div>

'''

# Generate company data for heat map
industry_data = []
for industry, companies in companies_by_industry.items():
    if companies:
        industry_data.append({
            'industry': industry,
            'count': len(companies),
            'companies': companies
        })
industry_data.sort(key=lambda x: x['count'], reverse=True)

# Generate companies HTML as heat map grid
companies_html = '''            <div class="heatmap-grid">
'''
for item in industry_data:
    industry = item['industry']
    count = item['count']
    companies = item['companies']
    
    # Color intensity based on count
    if count >= 20:
        intensity_class = 'heat-very-high'
    elif count >= 15:
        intensity_class = 'heat-high'
    elif count >= 10:
        intensity_class = 'heat-medium'
    elif count >= 5:
        intensity_class = 'heat-low'
    else:
        intensity_class = 'heat-very-low'
    
    companies_html += f'''                <div class="heat-cell {intensity_class}">
                    <div class="heat-header">
                        <div class="heat-industry">{industry}</div>
                        <div class="heat-count">{count}</div>
                    </div>
                    <details class="heat-details">
                        <summary>{count} companies</summary>
                        <div class="company-list">
'''
    for comp in companies:
        companies_html += f'                            <div class="company-item">• {comp}</div>\n'
    companies_html += '''                        </div>
                    </details>
                </div>
'''
companies_html += '''            </div>
'''

# Create complete HTML
html = f'''<!DOCTYPE html>
<html lang="en" data-theme="light">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Job Search Statistics | Melissa Witte-Spencer</title>
    <link rel="stylesheet" href="styles.css">
    <style>
        .stats-header {{
            background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
            color: white;
            padding: 40px 30px;
            text-align: center;
        }}

        .stats-header h1 {{
            margin: 0 0 12px 0;
            font-size: 32px;
        }}

        .stats-header p {{
            margin: 0;
            font-size: 16px;
            opacity: 0.95;
        }}

        .stats-content {{
            padding: 30px;
            max-width: 1400px;
            margin: 0 auto;
        }}

        .key-metrics {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 50px;
        }}

        .metric-card {{
            background: var(--bg-white);
            padding: 24px;
            border-radius: 12px;
            text-align: center;
            border: 2px solid var(--border);
            box-shadow: var(--shadow-sm);
            transition: var(--transition);
        }}

        .metric-card:hover {{
            transform: translateY(-4px);
            box-shadow: var(--shadow-md);
        }}

        .metric-value {{
            font-size: 48px;
            font-weight: 700;
            background: linear-gradient(135deg, var(--primary), var(--accent));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 8px;
        }}

        .metric-label {{
            font-size: 14px;
            color: var(--text-light);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            font-weight: 600;
        }}

        .section-title {{
            font-size: 26px;
            font-weight: 700;
            color: var(--text);
            margin: 50px 0 24px 0;
            display: flex;
            align-items: center;
            gap: 12px;
            padding-bottom: 12px;
            border-bottom: 3px solid var(--primary);
        }}

        /* Timeline - Monthly View */
        .timeline-chart {{
            background: var(--bg-white);
            padding: 30px;
            border-radius: 12px;
            border: 2px solid var(--border);
            margin-bottom: 50px;
        }}

        .timeline-bar {{
            display: flex;
            align-items: center;
            gap: 16px;
            margin-bottom: 16px;
        }}

        .timeline-month {{
            font-weight: 700;
            color: var(--text);
            font-size: 14px;
            min-width: 100px;
        }}

        .timeline-visual {{
            flex: 1;
            height: 44px;
            background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
            border-radius: 8px;
            display: flex;
            align-items: center;
            padding: 0 16px;
            color: white;
            font-weight: 700;
            font-size: 15px;
            box-shadow: var(--shadow-sm);
            transition: var(--transition);
        }}

        .timeline-visual:hover {{
            transform: scale(1.02);
            box-shadow: var(--shadow-md);
        }}

        .timeline-count {{
            white-space: nowrap;
        }}

        /* Role Categories - Bar Chart Visualization */
        .category-visual {{
            background: var(--bg-white);
            padding: 24px;
            border-radius: 12px;
            border: 2px solid var(--border);
            margin-bottom: 20px;
            transition: var(--transition);
        }}

        .category-visual:hover {{
            box-shadow: var(--shadow-md);
            border-color: var(--primary);
        }}

        .category-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 12px;
        }}

        .category-name {{
            font-size: 18px;
            font-weight: 700;
            color: var(--text);
            margin: 0;
        }}

        .category-count {{
            font-size: 14px;
            font-weight: 600;
            color: var(--text-light);
            background: var(--bg-light);
            padding: 4px 12px;
            border-radius: 12px;
        }}

        .category-bar-container {{
            background: var(--bg-light);
            border-radius: 8px;
            height: 48px;
            position: relative;
            overflow: hidden;
            margin-bottom: 12px;
        }}

        .category-bar {{
            height: 100%;
            background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
            border-radius: 8px;
            display: flex;
            align-items: center;
            justify-content: flex-end;
            padding-right: 16px;
            transition: width 0.5s ease;
            position: relative;
            min-width: 60px;
        }}

        .bar-label {{
            color: white;
            font-weight: 700;
            font-size: 16px;
            text-shadow: 0 1px 2px rgba(0,0,0,0.2);
        }}

        .category-details {{
            margin-top: 12px;
        }}

        .category-details summary {{
            cursor: pointer;
            font-size: 13px;
            color: var(--primary);
            font-weight: 600;
            padding: 8px 0;
            user-select: none;
        }}

        .category-details summary:hover {{
            text-decoration: underline;
        }}

        .positions-list {{
            margin-top: 12px;
            padding: 16px;
            background: var(--bg-light);
            border-radius: 8px;
            max-height: 300px;
            overflow-y: auto;
        }}

        .position-item {{
            padding: 6px 0;
            font-size: 13px;
            color: var(--text);
            line-height: 1.5;
            border-bottom: 1px solid var(--border);
        }}

        .position-item:last-child {{
            border-bottom: none;
        }}

        /* Heat Map for Companies by Industry */
        .heatmap-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}

        .heat-cell {{
            border-radius: 12px;
            padding: 24px;
            border: 2px solid var(--border);
            transition: var(--transition);
            position: relative;
            overflow: hidden;
        }}

        .heat-cell::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
            background: linear-gradient(135deg, var(--primary), var(--accent));
        }}

        .heat-very-high {{
            background: linear-gradient(135deg, rgba(33, 128, 195, 0.25) 0%, rgba(50, 184, 198, 0.25) 100%);
            border-color: var(--primary);
        }}

        .heat-high {{
            background: linear-gradient(135deg, rgba(33, 128, 195, 0.18) 0%, rgba(50, 184, 198, 0.18) 100%);
            border-color: var(--primary);
        }}

        .heat-medium {{
            background: linear-gradient(135deg, rgba(33, 128, 195, 0.12) 0%, rgba(50, 184, 198, 0.12) 100%);
        }}

        .heat-low {{
            background: linear-gradient(135deg, rgba(33, 128, 195, 0.08) 0%, rgba(50, 184, 198, 0.08) 100%);
        }}

        .heat-very-low {{
            background: linear-gradient(135deg, rgba(33, 128, 195, 0.04) 0%, rgba(50, 184, 198, 0.04) 100%);
        }}

        .heat-cell:hover {{
            transform: translateY(-4px);
            box-shadow: var(--shadow-md);
        }}

        .heat-header {{
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 12px;
        }}

        .heat-industry {{
            font-size: 16px;
            font-weight: 700;
            color: var(--text);
            line-height: 1.3;
            flex: 1;
        }}

        .heat-count {{
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, var(--primary), var(--accent));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-left: 12px;
        }}

        .heat-details summary {{
            cursor: pointer;
            font-size: 13px;
            color: var(--text-light);
            font-weight: 600;
            padding: 8px 0;
            user-select: none;
        }}

        .heat-details summary:hover {{
            color: var(--primary);
        }}

        .company-list {{
            margin-top: 12px;
            padding: 12px;
            background: rgba(255, 255, 255, 0.6);
            border-radius: 8px;
            max-height: 250px;
            overflow-y: auto;
        }}

        .company-item {{
            padding: 5px 0;
            font-size: 12px;
            color: var(--text);
            line-height: 1.4;
        }}

        @media (max-width: 768px) {{
            .heatmap-grid {{
                grid-template-columns: 1fr;
            }}
            
            .timeline-bar {{
                flex-direction: column;
                align-items: stretch;
                gap: 8px;
            }}
            
            .timeline-month {{
                min-width: auto;
            }}

            .category-header {{
                flex-direction: column;
                align-items: flex-start;
                gap: 8px;
            }}

            .heat-count {{
                font-size: 24px;
            }}
        }}
    </style>
</head>
<body>
    <a href="#main" class="skip-link">Skip to main content</a>
    
    <!-- Dark Mode Toggle -->
    <button class="theme-toggle" onclick="toggleTheme()" aria-label="Toggle dark mode">
        <span class="theme-toggle-icon" id="themeIcon">🌙</span>
        <span id="themeLabel">Dark</span>
    </button>

    <header role="banner">
        <div class="header-content">
            <div>
                <div class="logo"><a href="index.html">Melissa Witte-Spencer</a></div>
                <div class="logo-subtitle">Job Search Analytics & Strategic Targeting</div>
            </div>
        </div>
    </header>

    <!-- Hamburger Navigation -->
    <div class="floating-nav">
        <button class="hamburger-btn" onclick="toggleNavMenu()" aria-label="Navigation Menu" aria-expanded="false">
            <span></span>
            <span></span>
            <span></span>
        </button>
        <div class="nav-menu" id="navMenu" role="navigation">
            <button onclick="navigateTo('index.html')">🏠 Home</button>
            <button onclick="navigateTo('star-stories.html')">⭐ STAR Stories</button>
            <button onclick="navigateTo('iso-matrix.html')">📊 ISO Matrix</button>
            <button onclick="navigateTo('stats.html')" class="active">📈 Stats</button>
        </div>
    </div>

    <div class="stats-header">
        <h1>Job Search Activity Dashboard</h1>
        <p>Strategic targeting across high-growth technology companies | GRC Leadership Roles</p>
    </div>

    <div class="stats-content" id="main">
        <!-- Key Metrics -->
        <div class="key-metrics">
            <div class="metric-card">
                <div class="metric-value">{total_submissions}</div>
                <div class="metric-label">Total Applications</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{total_companies}</div>
                <div class="metric-label">Companies Targeted</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{total_title_groups}</div>
                <div class="metric-label">Title Categories</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">46</div>
                <div class="metric-label">Days Active</div>
            </div>
        </div>

        <!-- Application Timeline -->
        <h2 class="section-title">📈 Application Activity by Month</h2>
        <div class="timeline-chart">
{timeline_html}        </div>

        <!-- Position Titles by Role Similarity -->
        <h2 class="section-title">💼 Application Distribution by Role Category</h2>
        <p style="margin: -16px 0 24px 0; color: var(--text-light); font-size: 14px;">Horizontal bar chart showing position counts across 10 strategic role categories</p>
{positions_html}
        <!-- Companies by Industry -->
        <h2 class="section-title">🏢 Industry Coverage Heat Map</h2>
        <p style="margin: -16px 0 24px 0; color: var(--text-light); font-size: 14px;">Color intensity indicates company concentration across 8 industry sectors</p>
{companies_html}
    </div>

    <footer>
        <p>&copy; 2026 Melissa Witte-Spencer. All rights reserved. | Last Updated: January 10, 2026</p>
    </footer>

    <script>
        function toggleTheme() {{
            const html = document.documentElement;
            const currentTheme = html.getAttribute('data-theme');
            const newTheme = currentTheme === 'light' ? 'dark' : 'light';
            html.setAttribute('data-theme', newTheme);
            
            const icon = document.getElementById('themeIcon');
            const label = document.getElementById('themeLabel');
            if (newTheme === 'dark') {{
                icon.textContent = '☀️';
                label.textContent = 'Light';
            }} else {{
                icon.textContent = '🌙';
                label.textContent = 'Dark';
            }}
            
            localStorage.setItem('theme', newTheme);
        }}

        function toggleNavMenu() {{
            const menu = document.getElementById('navMenu');
            const hamburger = document.querySelector('.hamburger-btn');
            menu.classList.toggle('active');
            hamburger.classList.toggle('active');
        }}

        function navigateTo(page) {{
            const menu = document.getElementById('navMenu');
            menu.classList.remove('active');
            document.querySelector('.hamburger-btn').classList.remove('active');
            setTimeout(() => {{
                window.location.href = page;
            }}, 100);
        }}

        // Apply saved theme on load
        document.addEventListener('DOMContentLoaded', () => {{
            const savedTheme = localStorage.getItem('theme') || 'light';
            document.documentElement.setAttribute('data-theme', savedTheme);
            const icon = document.getElementById('themeIcon');
            const label = document.getElementById('themeLabel');
            if (savedTheme === 'dark') {{
                icon.textContent = '☀️';
                label.textContent = 'Light';
            }}
        }});
    </script>
</body>
</html>
'''

# Write the file
with open('stats.html', 'w') as f:
    f.write(html)

print(f"✅ Generated optimized stats.html")
print(f"   Total Submissions: {total_submissions}")
print(f"   Companies: {total_companies}")
print(f"   Title Categories: {total_title_groups}")
print(f"   Months: {len(sorted_months)}")
